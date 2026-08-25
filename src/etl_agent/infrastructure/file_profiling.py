"""文件格式识别、大小限制、摘要和脱敏 Profile 工具。"""

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import PurePath
from tempfile import SpooledTemporaryFile
from typing import Any, BinaryIO, cast

from openpyxl import load_workbook
from pyarrow import parquet as pq

_MAX_PROFILE_BYTES = 8 * 1024 * 1024
_SUPPORTED_FORMATS = {"csv", "json", "xlsx", "parquet"}
_SENSITIVE_MARKERS = (
    "password",
    "passwd",
    "secret",
    "token",
    "api_key",
    "email",
    "phone",
    "mobile",
    "身份证",
    "手机号",
)


class FileProfileError(ValueError):
    """表示文件格式、大小或 Profile 解析失败。"""


@dataclass
class FileInspection:
    """文件对象上传前的摘要和脱敏 Profile。"""

    fileobj: SpooledTemporaryFile[bytes]
    original_filename: str
    content_type: str
    file_format: str
    size_bytes: int
    sha256: str
    schema_snapshot: dict[str, Any]


def inspect_upload(
    source: BinaryIO,
    filename: str,
    content_type: str | None,
    max_size_bytes: int,
) -> FileInspection:
    """流式计量文件并生成受大小限制的格式 Profile，源数据不写入数据库。"""
    safe_filename = PurePath(filename or "upload").name
    file_format = _detect_format(safe_filename)
    if file_format not in _SUPPORTED_FORMATS:
        raise FileProfileError("仅支持 CSV、JSON、XLSX 和 Parquet 文件")
    temp = SpooledTemporaryFile(max_size=min(max_size_bytes, _MAX_PROFILE_BYTES), mode="w+b")
    digest = hashlib.sha256()
    size = 0
    try:
        while chunk := source.read(1024 * 1024):
            size += len(chunk)
            if size > max_size_bytes:
                raise FileProfileError("文件超过上传大小限制")
            digest.update(chunk)
            temp.write(chunk)
        temp.seek(0)
        schema_snapshot = _profile_file(cast(BinaryIO, temp), file_format)
        temp.seek(0)
        return FileInspection(
            fileobj=temp,
            original_filename=safe_filename[:512],
            content_type=content_type or "application/octet-stream",
            file_format=file_format,
            size_bytes=size,
            sha256=digest.hexdigest(),
            schema_snapshot=schema_snapshot,
        )
    except Exception:
        temp.close()
        raise


def _detect_format(filename: str) -> str:
    """根据安全化文件名后缀识别允许的文件格式。"""
    return PurePath(filename).suffix.lower().lstrip(".")


def _profile_file(fileobj: BinaryIO, file_format: str) -> dict[str, Any]:
    """按文件格式读取有限样本并返回字段推断结果。"""
    if file_format == "csv":
        return _profile_rows(_read_csv_rows(fileobj))
    if file_format == "json":
        return _profile_rows(_read_json_rows(fileobj))
    if file_format == "xlsx":
        return _profile_rows(_read_xlsx_rows(fileobj))
    if file_format == "parquet":
        return _profile_parquet(fileobj)
    raise FileProfileError("文件格式不受支持")


def _read_csv_rows(fileobj: BinaryIO) -> list[dict[str, Any]]:
    """读取 CSV 前 100 行并以首行作为列名。"""
    raw = fileobj.read(_MAX_PROFILE_BYTES)
    text = raw.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))[:101]
    if not rows:
        return []
    headers = _unique_headers(rows[0])
    return [dict(zip(headers, row, strict=False)) for row in rows[1:]]


def _read_json_rows(fileobj: BinaryIO) -> list[dict[str, Any]]:
    """读取小型 JSON 对象或对象数组，超出 Profile 预算时返回空样本。"""
    raw = fileobj.read(_MAX_PROFILE_BYTES)
    try:
        data = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return []
    if isinstance(data, dict):
        data = data.get("data", [data])
    if not isinstance(data, list):
        return []
    return [item for item in data[:100] if isinstance(item, dict)]


def _read_xlsx_rows(fileobj: BinaryIO) -> list[dict[str, Any]]:
    """读取 XLSX 第一个工作表前 100 行。"""
    workbook = load_workbook(fileobj, read_only=True, data_only=True)
    try:
        rows = list(next(iter(workbook.worksheets)).iter_rows(values_only=True))[:101]
    finally:
        workbook.close()
    if not rows:
        return []
    headers = _unique_headers([str(value or "") for value in rows[0]])
    return [dict(zip(headers, row, strict=False)) for row in rows[1:]]


def _profile_parquet(fileobj: BinaryIO) -> dict[str, Any]:
    """读取 Parquet Schema 和前 100 行的脱敏样本。"""
    parquet_file = pq.ParquetFile(fileobj)
    schema = [
        {"name": field.name, "inferred_type": str(field.type), "nullable": True}
        for field in parquet_file.schema_arrow
    ]
    table = parquet_file.read_row_groups([0]).slice(0, 100) if parquet_file.num_row_groups else None
    rows = table.to_pylist() if table is not None else []
    return {
        "version": "v1",
        "format": "parquet",
        "columns": schema,
        "sample_rows": _redact_rows(rows),
    }


def _profile_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """从结构化行样本推断列类型并生成脱敏样本。"""
    names = list(dict.fromkeys(key for row in rows for key in row))
    columns = [
        {
            "name": name,
            "inferred_type": _infer_type([row.get(name) for row in rows]),
            "nullable": any(row.get(name) in (None, "") for row in rows),
        }
        for name in names
    ]
    return {
        "version": "v1",
        "columns": columns,
        "sample_rows": _redact_rows(rows[:100]),
    }


def _unique_headers(headers: list[str]) -> list[str]:
    """清洗空列名并为重复列名追加稳定序号。"""
    counts: dict[str, int] = {}
    result: list[str] = []
    for index, value in enumerate(headers, start=1):
        base = value.strip() or f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def _infer_type(values: list[Any]) -> str:
    """根据有限样本按整数、浮点、布尔、日期和字符串顺序推断类型。"""
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return "unknown"
    if all(
        isinstance(value, bool) or str(value).lower() in {"true", "false"} for value in non_empty
    ):
        return "boolean"
    if all(_is_integer(value) for value in non_empty):
        return "integer"
    if all(_is_float(value) for value in non_empty):
        return "number"
    if all(_is_date(value) for value in non_empty):
        return "datetime"
    return "string"


def _is_integer(value: Any) -> bool:
    """判断样本值是否可解释为整数。"""
    return (
        isinstance(value, int)
        and not isinstance(value, bool)
        or bool(re.fullmatch(r"[-+]?\d+", str(value)))
    )


def _is_float(value: Any) -> bool:
    """判断样本值是否可解释为浮点数。"""
    try:
        float(value)
        return not isinstance(value, bool)
    except (TypeError, ValueError):
        return False


def _is_date(value: Any) -> bool:
    """判断样本值是否为日期时间或常见 ISO 日期文本。"""
    if isinstance(value, (date, datetime)):
        return True
    try:
        datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return "-" in str(value)
    except ValueError:
        return False


def _redact_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按列名脱敏并转换为 JSON 可序列化的样本行。"""
    return [
        {str(key): _redact_value(str(key), value) for key, value in row.items()} for row in rows
    ]


def _redact_value(field_name: str, value: Any) -> Any:
    """隐藏敏感字段并限制普通文本长度。"""
    normalized = field_name.lower().replace("-", "_")
    if any(marker in normalized for marker in _SENSITIVE_MARKERS):
        return "[REDACTED]"
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (datetime, date, Decimal)):
        return str(value)
    if isinstance(value, bytes):
        return f"[BYTES:{len(value)}]"
    return str(value)[:256]
